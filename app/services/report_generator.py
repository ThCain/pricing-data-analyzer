import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from typing import Dict, Any, List
import logging
from datetime import datetime
import io
import base64

logger = logging.getLogger(__name__)

class ReportGenerator:
    def __init__(self):
        self.reports_dir = "reports"
        self.visualizations_dir = "visualizations"
        
        # Create directories if they don't exist
        os.makedirs(self.reports_dir, exist_ok=True)
        os.makedirs(self.visualizations_dir, exist_ok=True)
        
        # Set up matplotlib style
        plt.style.use('seaborn-v0_8')
        sns.set_palette("husl")
    
    def generate_excel_report(self, df: pd.DataFrame, file_id: str, original_filename: str) -> str:
        """Generate comprehensive Excel report"""
        try:
            report_path = os.path.join(self.reports_dir, f"pricing_report_{file_id}.xlsx")
            
            # Create Excel workbook
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create sheets
            summary_sheet = wb.create_sheet("Summary", 0)
            data_sheet = wb.create_sheet("Data", 1)
            stats_sheet = wb.create_sheet("Statistics", 2)
            outliers_sheet = wb.create_sheet("Outliers", 3)
            trends_sheet = wb.create_sheet("Trends", 4)
            
            # Generate summary sheet
            self._create_summary_sheet(summary_sheet, df, original_filename)
            
            # Generate data sheet
            self._create_data_sheet(data_sheet, df)
            
            # Generate statistics sheet
            self._create_statistics_sheet(stats_sheet, df)
            
            # Generate outliers sheet
            self._create_outliers_sheet(outliers_sheet, df)
            
            # Generate trends sheet
            self._create_trends_sheet(trends_sheet, df)
            
            # Add charts to summary sheet
            self._add_charts_to_summary(summary_sheet, df)
            
            # Save workbook
            wb.save(report_path)
            
            logger.info(f"Excel report generated: {report_path}")
            return report_path
            
        except Exception as e:
            logger.error(f"Error generating Excel report: {str(e)}")
            raise
    
    def _create_summary_sheet(self, sheet, df: pd.DataFrame, original_filename: str):
        """Create summary sheet with key metrics"""
        # Title
        sheet['A1'] = "Pricing Data Analysis Report"
        sheet['A1'].font = Font(size=16, bold=True)
        sheet['A1'].alignment = Alignment(horizontal='center')
        sheet.merge_cells('A1:D1')
        
        # File info
        sheet['A3'] = f"Original Filename: {original_filename}"
        sheet['A4'] = f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        sheet['A5'] = f"Total Records: {len(df):,}"
        
        # Key metrics
        basic_stats = df['price'].describe()
        
        row = 8
        sheet[f'A{row}'] = "Key Metrics"
        sheet[f'A{row}'].font = Font(size=14, bold=True)
        row += 2
        
        metrics = [
            ("Average Price", f"${basic_stats['mean']:.2f}"),
            ("Median Price", f"${basic_stats['50%']:.2f}"),
            ("Minimum Price", f"${basic_stats['min']:.2f}"),
            ("Maximum Price", f"${basic_stats['max']:.2f}"),
            ("Price Range", f"${basic_stats['max'] - basic_stats['min']:.2f}"),
            ("Standard Deviation", f"${basic_stats['std']:.2f}"),
            ("Date Range", f"{df['date'].min().date()} to {df['date'].max().date()}"),
            ("Unique Products", f"{df['product_name'].nunique():,}"),
            ("Unique Categories", f"{df['category'].nunique() if 'category' in df.columns else 0:,}")
        ]
        
        for metric, value in metrics:
            sheet[f'A{row}'] = metric
            sheet[f'B{row}'] = value
            sheet[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Apply styling
        self._apply_sheet_styling(sheet)
    
    def _create_data_sheet(self, sheet, df: pd.DataFrame):
        """Create sheet with raw data"""
        # Write data
        for r in dataframe_to_rows(df, index=False, header=True):
            sheet.append(r)
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Style header
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
    
    def _create_statistics_sheet(self, sheet, df: pd.DataFrame):
        """Create detailed statistics sheet"""
        # Title
        sheet['A1'] = "Detailed Statistics"
        sheet['A1'].font = Font(size=14, bold=True)
        
        # Price statistics
        row = 3
        sheet[f'A{row}'] = "Price Statistics"
        sheet[f'A{row}'].font = Font(size=12, bold=True)
        row += 1
        
        price_stats = df['price'].describe()
        for stat, value in price_stats.items():
            sheet[f'A{row}'] = stat.replace('_', ' ').title()
            sheet[f'B{row}'] = f"${value:.2f}"
            row += 1
        
        # Category statistics
        if 'category' in df.columns:
            row += 2
            sheet[f'A{row}'] = "Category Statistics"
            sheet[f'A{row}'].font = Font(size=12, bold=True)
            row += 1
            
            category_stats = df.groupby('category')['price'].agg(['count', 'mean', 'min', 'max']).round(2)
            category_stats.columns = ['Count', 'Average', 'Min', 'Max']
            
            # Headers
            sheet[f'A{row}'] = "Category"
            sheet[f'B{row}'] = "Count"
            sheet[f'C{row}'] = "Average Price"
            sheet[f'D{row}'] = "Min Price"
            sheet[f'E{row}'] = "Max Price"
            
            for cell in sheet[row:row][0:5]:
                cell.font = Font(bold=True)
            
            row += 1
            
            for category, stats in category_stats.iterrows():
                sheet[f'A{row}'] = category
                sheet[f'B{row}'] = stats['Count']
                sheet[f'C{row}'] = f"${stats['Average']:.2f}"
                sheet[f'D{row}'] = f"${stats['Min']:.2f}"
                sheet[f'E{row}'] = f"${stats['Max']:.2f}"
                row += 1
    
    def _create_outliers_sheet(self, sheet, df: pd.DataFrame):
        """Create outliers analysis sheet"""
        # Title
        sheet['A1'] = "Outliers Analysis"
        sheet['A1'].font = Font(size=14, bold=True)
        
        # Calculate outliers using IQR method
        q1, q3 = df['price'].quantile([0.25, 0.75])
        iqr = q3 - q1
        lower_bound = q1 - 1.5 * iqr
        upper_bound = q3 + 1.5 * iqr
        
        outliers = df[(df['price'] < lower_bound) | (df['price'] > upper_bound)]
        
        row = 3
        sheet[f'A{row}'] = f"Total Outliers Detected: {len(outliers)}"
        sheet[f'A{row}'].font = Font(bold=True)
        row += 1
        sheet[f'A{row}'] = f"Percentage of Data: {(len(outliers)/len(df)*100):.2f}%"
        row += 2
        
        if len(outliers) > 0:
            # Outlier details
            sheet[f'A{row}'] = "Outlier Details"
            sheet[f'A{row}'].font = Font(size=12, bold=True)
            row += 1
            
            # Headers
            headers = ['Product Name', 'Price', 'Date', 'Category', 'Supplier']
            for i, header in enumerate(headers):
                sheet.cell(row=row, column=i+1, value=header).font = Font(bold=True)
            row += 1
            
            # Outlier data
            for _, outlier in outliers.head(50).iterrows():  # Limit to first 50
                sheet.cell(row=row, column=1, value=outlier['product_name'])
                sheet.cell(row=row, column=2, value=f"${outlier['price']:.2f}")
                sheet.cell(row=row, column=3, value=outlier['date'].date())
                sheet.cell(row=row, column=4, value=outlier.get('category', 'N/A'))
                sheet.cell(row=row, column=5, value=outlier.get('supplier', 'N/A'))
                row += 1
        else:
            sheet[f'A{row}'] = "No outliers detected in the dataset."
    
    def _create_trends_sheet(self, sheet, df: pd.DataFrame):
        """Create trends analysis sheet"""
        # Title
        sheet['A1'] = "Trends Analysis"
        sheet['A1'].font = Font(size=14, bold=True)
        
        # Monthly trends
        df_sorted = df.sort_values('date')
        monthly_trends = df_sorted.groupby(df_sorted['date'].dt.to_period('M')).agg({
            'price': ['mean', 'count']
        }).round(2)
        
        monthly_trends.columns = ['Average Price', 'Transaction Count']
        
        row = 3
        sheet[f'A{row}'] = "Monthly Price Trends"
        sheet[f'A{row}'].font = Font(size=12, bold=True)
        row += 1
        
        # Headers
        sheet[f'A{row}'] = "Month"
        sheet[f'B{row}'] = "Average Price"
        sheet[f'C{row}'] = "Transaction Count"
        
        for cell in sheet[row:row][0:3]:
            cell.font = Font(bold=True)
        
        row += 1
        
        # Monthly data
        for month, stats in monthly_trends.iterrows():
            sheet[f'A{row}'] = str(month)
            sheet[f'B{row}'] = f"${stats['Average Price']:.2f}"
            sheet[f'C{row}'] = stats['Transaction Count']
            row += 1
    
    def _add_charts_to_summary(self, sheet, df: pd.DataFrame):
        """Add charts to the summary sheet"""
        try:
            # Price distribution chart
            price_dist = df['price'].value_counts(bins=10).sort_index()
            
            # Create bar chart for price distribution
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "Price Distribution"
            chart.y_axis.title = "Frequency"
            chart.x_axis.title = "Price Range"
            
            # Data for chart (simplified for example)
            data = Reference(sheet, min_col=2, min_row=8, max_col=2, max_row=17)
            categories = Reference(sheet, min_col=1, min_row=8, max_row=17)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            
            sheet.add_chart(chart, "G2")
            
        except Exception as e:
            logger.warning(f"Could not add charts to summary sheet: {str(e)}")
    
    def _apply_sheet_styling(self, sheet):
        """Apply consistent styling to sheet"""
        # Set column widths
        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['D'].width = 15
        
        # Add borders to data area
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in sheet.iter_rows(min_row=8, max_row=20, min_col=1, max_col=2):
            for cell in row:
                if cell.value:
                    cell.border = thin_border
    
    def generate_visualizations(self, df: pd.DataFrame, file_id: str) -> Dict[str, str]:
        """Generate data visualizations and return paths"""
        visualization_paths = {}
        
        try:
            # Price distribution histogram
            plt.figure(figsize=(10, 6))
            plt.hist(df['price'], bins=30, alpha=0.7, edgecolor='black')
            plt.title('Price Distribution')
            plt.xlabel('Price ($)')
            plt.ylabel('Frequency')
            plt.grid(True, alpha=0.3)
            
            hist_path = os.path.join(self.visualizations_dir, f"price_dist_{file_id}.png")
            plt.savefig(hist_path, dpi=300, bbox_inches='tight')
            plt.close()
            visualization_paths['price_distribution'] = hist_path
            
            # Price over time
            plt.figure(figsize=(12, 6))
            df_sorted = df.sort_values('date')
            plt.plot(df_sorted['date'], df_sorted['price'], alpha=0.6)
            plt.title('Price Trends Over Time')
            plt.xlabel('Date')
            plt.ylabel('Price ($)')
            plt.xticks(rotation=45)
            plt.grid(True, alpha=0.3)
            
            trend_path = os.path.join(self.visualizations_dir, f"price_trend_{file_id}.png")
            plt.savefig(trend_path, dpi=300, bbox_inches='tight')
            plt.close()
            visualization_paths['price_trend'] = trend_path
            
            # Box plot by category
            if 'category' in df.columns and df['category'].nunique() <= 10:
                plt.figure(figsize=(12, 8))
                df.boxplot(column='price', by='category', ax=plt.gca())
                plt.title('Price Distribution by Category')
                plt.xlabel('Category')
                plt.ylabel('Price ($)')
                plt.xticks(rotation=45)
                
                category_path = os.path.join(self.visualizations_dir, f"price_by_category_{file_id}.png")
                plt.savefig(category_path, dpi=300, bbox_inches='tight')
                plt.close()
                visualization_paths['price_by_category'] = category_path
            
            # Monthly average prices
            plt.figure(figsize=(10, 6))
            monthly_avg = df.groupby(df['date'].dt.to_period('M'))['price'].mean()
            monthly_avg.plot(kind='bar')
            plt.title('Monthly Average Prices')
            plt.xlabel('Month')
            plt.ylabel('Average Price ($)')
            plt.xticks(rotation=45)
            plt.grid(True, alpha=0.3)
            
            monthly_path = os.path.join(self.visualizations_dir, f"monthly_avg_{file_id}.png")
            plt.savefig(monthly_path, dpi=300, bbox_inches='tight')
            plt.close()
            visualization_paths['monthly_average'] = monthly_path
            
            logger.info(f"Generated {len(visualization_paths)} visualizations for file {file_id}")
            
        except Exception as e:
            logger.error(f"Error generating visualizations: {str(e)}")
        
        return visualization_paths
    
    def get_base64_image(self, image_path: str) -> str:
        """Convert image to base64 for embedding in reports"""
        try:
            with open(image_path, "rb") as image_file:
                return base64.b64encode(image_file.read()).decode('utf-8')
        except Exception as e:
            logger.error(f"Error converting image to base64: {str(e)}")
            return ""
